from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.contrib.auth import authenticate
from django.utils import timezone
from datetime import timedelta
import openpyxl

from .models import User, Rango
from .serializers import UserSerializer, RangoSerializer, UserListSerializer


class IsStaff(IsAuthenticated):
    """Permiso: solo usuarios staff (admin)."""
    def has_permission(self, request, view):
        return super().has_permission(request, view) and request.user.is_staff


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username') # Que servirá para usu_cod o username
        dni = request.data.get('dni')           # usu_dni

        if not username or not dni:
            return Response({'error': 'Username/UsuCod and DNI are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Autenticación estándar para todos los usuarios
        user = authenticate(request, username=username, password=dni)
        if user:
            serializer = UserSerializer(user)
            return Response(serializer.data)
        else:
            return Response({'error': 'Credenciales inválidas'}, status=status.HTTP_401_UNAUTHORIZED)


class UserScoreView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            # Soportando pk string por usu_cod o id int
            if str(pk).isdigit():
                user = User.objects.get(pk=pk)
            else:
                user = User.objects.get(usu_cod=pk)
            serializer = UserSerializer(user)
            return Response(serializer.data)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)


class UserListView(APIView):
    """Lista todos los usuarios (solo staff)."""
    permission_classes = [IsStaff]

    def get(self, request):
        users = User.objects.all().order_by('-usu_pun_tot')
        serializer = UserListSerializer(users, many=True)
        return Response(serializer.data)


class ImportUsersView(APIView):
    """
    Importa usuarios desde un archivo Excel (.xlsx).
    Hace UPSERT por DNI (usu_dni):
    """
    permission_classes = [IsStaff]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No se proporcionó ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file_obj.name.endswith('.xlsx'):
            return Response({'error': 'El archivo debe ser formato .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'error': 'No se pudo leer el archivo Excel. Verifica el formato.'}, status=status.HTTP_400_BAD_REQUEST)

        header_row = [str(cell.value).strip().upper() if cell.value else '' for cell in ws[1]]
        
        col_map = {
            'username': 'ASESOR ANALISTA',
            'dni': 'NUMERO DOCUMENTO',
            'nombre_completo': 'NOMBRE COMPLETO',
            'genero': 'GENERO',
            'edad': 'EDAD',
            'zona': 'ZONA',
            'antiguedad': 'ANTIGUEDAD',
            'categoria': 'CATEGORIA'
        }

        required_cols = {col_map['username'], col_map['dni']}
        if not required_cols.issubset(set(header_row)):
            missing = required_cols - set(header_row)
            return Response(
                {'error': f'Faltan columnas requeridas: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        idx = {key: header_row.index(val) if val in header_row else None for key, val in col_map.items()}
        
        creados = 0
        actualizados = 0
        errores = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            try:
                # Extraer y limpiar valores de forma segura
                def get_val(key, default=''):
                    v = row[idx[key]] if idx[key] is not None else None
                    if v is None: return default
                    return str(v).strip()

                username_val = get_val('username')
                dni_val = get_val('dni')
                nombre_val = get_val('nombre_completo')
                genero_val = get_val('genero').upper() if get_val('genero') else None
                
                try:
                    edad_val = int(row[idx['edad']]) if idx['edad'] is not None and row[idx['edad']] is not None else 0
                except (ValueError, TypeError): edad_val = 0
                
                zona_val = get_val('zona', None)
                
                try:
                    age_val = int(row[idx['antiguedad']]) if idx['antiguedad'] is not None and row[idx['antiguedad']] is not None else 0
                except (ValueError, TypeError): age_val = 0
                
                cat_val = get_val('categoria', None)

                if not dni_val or not username_val:
                    errores.append({'fila': row_idx, 'motivo': 'DNI o username vacío'})
                    continue
                
                # Mapear Género
                usu_sex = None
                if genero_val:
                    if 'MASC' in genero_val: usu_sex = 'M'
                    elif 'FEM' in genero_val: usu_sex = 'F'
                
                # Mapear Categoría (Búsqueda por CÓDIGO)
                cat_obj = None
                if cat_val:
                    from .models import Categoria
                    cat_obj = Categoria.objects.filter(cat_cod=cat_val).first()

                defaults = {
                    'username': username_val,
                    'usu_cod': username_val,
                    'usu_nom': nombre_val,
                    'usu_sex': usu_sex,
                    'usu_edad': edad_val,
                    'usu_zon': zona_val,
                    'usu_age': age_val,
                    'cat_cod': cat_obj
                }

                user, created = User.objects.update_or_create(
                    usu_dni=dni_val,
                    defaults=defaults
                )
                
                if created:
                    user.set_password(dni_val)
                    user.save()
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append({'fila': row_idx, 'motivo': str(e)})

        return Response({
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'total_procesados': creados + actualizados,
        }, status=status.HTTP_200_OK)


class ExportUserTemplateView(APIView):
    """
    Descarga una plantilla de Excel para importar usuarios.
    """
    permission_classes = [IsStaff]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Usuarios"

        headers = [
            'ASESOR ANALISTA', 'NUMERO DOCUMENTO', 'NOMBRE COMPLETO', 
            'GENERO', 'EDAD', 'ZONA', 'ANTIGUEDAD', 'CATEGORIA'
        ]
        ws.append(headers)

        # Ejemplo opcional
        ws.append(['jperez', '12345678', 'JUAN PEREZ', 'MASCULINO', 25, 'ZONA SUR', 2, '0'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_usuarios.xlsx'
        return response


class CleanupInactiveUsersView(APIView):
    """
    Limpia usuarios inactivos.
    """
    permission_classes = [IsStaff]

    def post(self, request):
        months = int(request.data.get('months', 2))
        delete_users = str(request.data.get('delete', 'false')).lower() == 'true'

        cutoff_date = timezone.now() - timedelta(days=months * 30)

        inactive_qs = User.objects.exclude(
            usu_cod='PEOPLEADMIN'
        ).filter(
            last_login__lt=cutoff_date
        ) | User.objects.exclude(usu_cod='PEOPLEADMIN').filter(last_login__isnull=True)

        inactive_qs = inactive_qs.exclude(pk=request.user.pk).distinct()
        count = inactive_qs.count()

        if delete_users:
            inactive_qs.delete()
            return Response({
                'accion': 'eliminados',
                'cantidad': count,
                'mensaje': f'{count} usuario(s) eliminado(s)'
            })
        else:
            updated = inactive_qs.update(usu_pun_tot=0, ran_sig=None)
            return Response({
                'accion': 'reset_puntos',
                'cantidad': updated,
                'mensaje': f'{updated} usuario(s) reseteado(s)'
            })


class RangoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Rango.objects.all().order_by('ran_pun_min')
    serializer_class = RangoSerializer
    permission_classes = [IsAuthenticated]


class RankingView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = User.objects.all().order_by('-usu_pun_tot')
        total_users = users.count()
        
        user_ids = list(users.values_list('pk', flat=True))
        try:
            user_rank = user_ids.index(request.user.pk) + 1
        except ValueError:
            user_rank = 0

        top_13 = users[:13]
        serializer = UserListSerializer(top_13, many=True)
        
        return Response({
            'top_users': serializer.data,
            'user_rank': user_rank,
            'total_users': total_users
        })
