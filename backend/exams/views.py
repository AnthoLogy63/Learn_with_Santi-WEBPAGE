import random
import openpyxl
import json
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import transaction
from django.utils import timezone
from django.db import models
from django.contrib.auth import get_user_model
from .models import TipoPregunta, Examen, CategoriaExamen, Pregunta, Opcion, Intento, Respuesta, PreguntaCompetencia, ExamenCategoriaCompetencia
from users.models import Categoria, Competencia
from .serializers import (
    ExamenSerializer, PreguntaSerializer, IntentoSerializer, 
    RespuestaSerializer, OpcionSerializer, TipoPreguntaSerializer
)


class IsStaff(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and request.user.is_staff

UserModel = get_user_model()


class ImportExamView(APIView):
    """
    Importa preguntas a un Examen desde Excel.
    Formato esperado:
    COLUMNAS: CATEGORIA, COMPETENCIA, PREGUNTA, PUNTOS, TIEMPO, OPCION_1, OPCION_2, OPCION_3, OPCION_4, CORRECTA (1-4)
    """
    permission_classes = [IsStaff]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file_obj = request.FILES.get('file')
        exa_cod_param = request.data.get('exa_cod') # Opcional, si queremos atarlo a uno ya existente
        
        if not file_obj:
            return Response({'error': 'No se proporcionó ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active
        except Exception:
            return Response({'error': 'No se pudo leer el archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        header_row = [str(cell.value).strip().upper() if cell.value else '' for cell in ws[1]]
        
        # Mapeo de columnas
        col_map = {
            'categoria': 'CATEGORIA',
            'competencia': 'COMPETENCIA',
            'pregunta': 'PREGUNTA',
            'puntos': 'PUNTOS',
            'tiempo': 'TIEMPO',
            'opc1': 'OPCION_1',
            'opc2': 'OPCION_2',
            'opc3': 'OPCION_3',
            'opc4': 'OPCION_4',
            'opc5': 'OPCION_5',
            'opc6': 'OPCION_6',
            'opc7': 'OPCION_7',
            'opc8': 'OPCION_8',
            'opc9': 'OPCION_9',
            'opc10': 'OPCION_10',
            'correcta': 'CORRECTA',
            'tipo': 'TIPO',
            'titulo_examen': 'EXAMEN'
        }

        # Verificación de columnas mínimas
        required = {col_map['pregunta'], col_map['opc1'], col_map['correcta']}
        if not required.issubset(set(header_row)):
            missing = required - set(header_row)
            return Response({'error': f'Faltan columnas requeridas: {", ".join(missing)}'}, status=status.HTTP_400_BAD_REQUEST)

        idx = {key: header_row.index(val) if val in header_row else None for key, val in col_map.items()}

        import uuid
        exam = None
        if exa_cod_param:
            exam = Examen.objects.filter(exa_cod=exa_cod_param).first()

        if not exam:
            # Intentar obtener el nombre del primer registro que tenga la columna EXAMEN
            exam_name = None
            if idx['titulo_examen'] is not None:
                for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                    if row[idx['titulo_examen']]:
                        exam_name = str(row[idx['titulo_examen']]).strip()
                        break
            
            if not exam_name:
                exam_name = file_obj.name.split('.')[0]

            exam = Examen.objects.create(
                exa_cod='EX_' + uuid.uuid4().hex[:8].upper(),
                exa_nom=exam_name,
                exa_des=f"Examen importado el {timezone.now().strftime('%d/%m/%Y')}"
            )
        else:
            # Si el examen ya existe pero viene un título en el excel, ¿lo actualizamos?
            # Por ahora solo si el admin no paso un exa_cod específico.
            pass

        creadas = 0
        errores = []

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                
                try:
                    # Extraer valores de forma segura
                    def get_val(key, default=''):
                        v = row[idx[key]] if idx[key] is not None else None
                        if v is None: return default
                        return str(v).strip()

                    cat_val = get_val('categoria', None)
                    com_val = get_val('competencia', None)
                    pre_val = get_val('pregunta')
                    
                    try: pun_val = int(row[idx['puntos']]) if idx['puntos'] is not None and row[idx['puntos']] else 10
                    except: pun_val = 10
                    
                    try: tie_val = int(row[idx['tiempo']]) if idx['tiempo'] is not None and row[idx['tiempo']] else 60
                    except: tie_val = 60
                    
                    try: tipo_val = int(row[idx['tipo']]) if idx['tipo'] is not None and row[idx['tipo']] else 1
                    except: tipo_val = 1
                    
                    correcta_idx = get_val('correcta')

                    if not pre_val: continue

                    # 1. Manejar Categoría vinculada al examen (Búsqueda por CÓDIGO)
                    if cat_val:
                        categoria = Categoria.objects.filter(cat_cod=cat_val).first()
                        if categoria:
                            CategoriaExamen.objects.get_or_create(exa_cod=exam, cat_cod=categoria)
                    else:
                        categoria = None

                    # 2. Manejar Competencia (Búsqueda por CÓDIGO)
                    if com_val:
                        competencia = Competencia.objects.filter(com_cod=com_val).first()
                        # Si hay categoría y competencia, asegurar config en ExamenCategoriaCompetencia
                        if categoria and competencia:
                            ExamenCategoriaCompetencia.objects.get_or_create(
                                exa_cod=exam, cat_cod=categoria, com_cod=competencia
                            )
                    else:
                        competencia = None

                    # 3. Crear Pregunta
                    p_cod = f"PRE_{uuid.uuid4().hex[:12].upper()}"
                    pregunta = Pregunta.objects.create(
                        pre_cod=p_cod,
                        exa_cod=exam,
                        pre_tex=pre_val,
                        pre_pun=pun_val,
                        pre_tie=tie_val,
                        tip_pre_cod_id=tipo_val
                    )

                    # 4. Vincular Pregunta a Competencia si existe
                    if competencia:
                        PreguntaCompetencia.objects.get_or_create(pre_cod=pregunta, com_cod=competencia)

                    # 5. Crear Opciones (Soportamos hasta 10 columnas para Relación)
                    for o_idx in range(1, 11):
                        key = f'opc{o_idx}'
                        if idx[key] is not None and row[idx[key]]:
                            Opcion.objects.create(
                                opc_cod=f"OPC_{uuid.uuid4().hex[:15].upper()}",
                                pre_cod=pregunta,
                                opc_tex=str(row[idx[key]]).strip(),
                                opc_cor=(str(o_idx) == correcta_idx)
                            )
                    
                    creadas += 1

                except Exception as e:
                    errores.append({'fila': row_idx, 'error': str(e)})

        return Response({
            'status': 'completado',
            'exa_cod': exam.exa_cod,
            'preguntas_creadas': creadas,
            'errores': errores
        }, status=status.HTTP_200_OK if not errores else status.HTTP_207_MULTI_STATUS)


class ExportExamTemplateView(APIView):
    """
    Descarga una plantilla de Excel para importar exámenes.
    """
    permission_classes = [IsStaff]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Examen"

        headers = [
            'EXAMEN', 'CATEGORIA', 'COMPETENCIA', 'PREGUNTA', 
            'PUNTOS', 'TIEMPO', 'TIPO', 
            'OPCION_1', 'OPCION_2', 'OPCION_3', 'OPCION_4', 
            'OPCION_5', 'OPCION_6', 'OPCION_7', 'OPCION_8', 
            'OPCION_9', 'OPCION_10', 'CORRECTA'
        ]
        ws.append(headers)

        # Ejemplo
        ws.append([
            'EXAMEN DE PRUEBA', '0', 'COMP-AGILIDAD-MENTAL', '¿Cuánto es 2+2?', 
            10, 60, 1, 
            '3', '4', '5', '6', 
            '', '', '', '', 
            '', '', '2'
        ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_examen.xlsx'
        return response


class ExamenViewSet(viewsets.ModelViewSet):
    queryset = Examen.objects.all()
    serializer_class = ExamenSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer(self, *args, **kwargs):
        # Si somos staff y estamos en 'list', usamos el modo simple
        if self.action == 'list' and self.request.user.is_staff:
            kwargs['simple'] = True
        return super().get_serializer(*args, **kwargs)

    def get_queryset(self):
        # A futuro: Filtrar los exámenes habilitados para la categoría del usuario
        if self.request.user.is_staff:
            return Examen.objects.all()
        # Mostrar exámenes asociados a la categoría del estudiante
        if self.request.user.cat_cod:
            return Examen.objects.filter(categoriaexamen__cat_cod=self.request.user.cat_cod)
        return Examen.objects.none()

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def categorias(self, request):
        from users.serializers import CategoriaSerializer
        cats = Categoria.objects.all()
        return Response(CategoriaSerializer(cats, many=True).data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def competencias(self, request):
        from users.serializers import CompetenciaSerializer
        comps = Competencia.objects.all()
        return Response(CompetenciaSerializer(comps, many=True).data)

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def export_excel(self, request, pk=None):
        """Exporta el examen a un formato Excel compatible con el importador."""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font

        exam = self.get_object()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Examen"

        # Headers
        headers = [
            'CATEGORIA', 'COMPETENCIA', 'PREGUNTA', 'PUNTOS', 'TIEMPO', 'TIPO',
            'OPCION_1', 'OPCION_2', 'OPCION_3', 'OPCION_4', 'OPCION_5', 
            'OPCION_6', 'OPCION_7', 'OPCION_8', 'OPCION_9', 'OPCION_10',
            'CORRECTA', 'EXAMEN'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Content
        row_num = 2
        preguntas = Pregunta.objects.filter(exa_cod=exam).prefetch_related('opciones')
        
        for pre in preguntas:
            cat_obj = CategoriaExamen.objects.filter(exa_cod=exam).first()
            cat_nom = cat_obj.cat_cod.cat_nom if cat_obj else ""
            comp_obj = PreguntaCompetencia.objects.filter(pre_cod=pre).first()
            comp_nom = comp_obj.com_cod.com_nom if comp_obj else ""

            ws.cell(row=row_num, column=1, value=cat_nom)
            ws.cell(row=row_num, column=2, value=comp_nom)
            ws.cell(row=row_num, column=3, value=pre.pre_tex)
            ws.cell(row=row_num, column=4, value=pre.pre_pun)
            ws.cell(row=row_num, column=5, value=pre.pre_tie)
            ws.cell(row=row_num, column=6, value=pre.tip_pre_cod_id or 1)
            
            correcta_idx = ""
            for i, opt in enumerate(pre.opciones.all()[:10], 1):
                ws.cell(row=row_num, column=6 + i, value=opt.opc_tex)
                if opt.opc_cor:
                    correcta_idx = str(i)
            
            ws.cell(row=row_num, column=17, value=correcta_idx)
            ws.cell(row=row_num, column=18, value=exam.exa_nom)
            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=examen_{exam.exa_cod}.xlsx'
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def start_or_resume(self, request, pk=None):
        """
        Endpoint 1 en 1: Devuelve el examen, preguntas y opciones generadas para este intento.
        """
        examen = self.get_object()
        user = request.user
        
        # Buscar intento en progreso
        intento = Intento.objects.filter(usu_cod=user, exa_cod=examen, exa_fec_fin__isnull=True).first()
        
        if not intento:
            # Crear Intento Nuevo
            num_intentos = Intento.objects.filter(usu_cod=user, exa_cod=examen).count()
            intento_cod = f"{user.usu_cod}_{examen.exa_cod}_{num_intentos+1}"
            
            intento = Intento.objects.create(
                int_cod=intento_cod,
                usu_cod=user,
                exa_cod=examen,
                exa_num_int=num_intentos + 1
            )

            # Generación de preguntas basada en la configuración ExamenCategoriaCompetencia
            preguntas_seleccionadas = []

            if user.cat_cod:
                configuraciones = ExamenCategoriaCompetencia.objects.filter(
                    exa_cod=examen, cat_cod=user.cat_cod
                )

                if configuraciones.exists():
                    # Una sola query: todas las preguntas del examen con sus competencias
                    preguntas_por_competencia: dict = {}
                    preg_comp_qs = PreguntaCompetencia.objects.filter(
                        pre_cod__exa_cod=examen
                    ).select_related('pre_cod').values('com_cod_id', 'pre_cod_id')

                    for pc in preg_comp_qs:
                        preguntas_por_competencia.setdefault(pc['com_cod_id'], []).append(pc['pre_cod_id'])

                    # Una sola query: todas las preguntas necesarias de una vez
                    todos_ids_necesarios = set()
                    for conf in configuraciones:
                        todos_ids_necesarios.update(preguntas_por_competencia.get(conf.com_cod_id, []))

                    preguntas_map = {
                        p.pre_cod: p
                        for p in Pregunta.objects.filter(pre_cod__in=todos_ids_necesarios)
                    }

                    for conf in configuraciones:
                        ids_comp = preguntas_por_competencia.get(conf.com_cod_id, [])
                        preguntas_disponibles = [preguntas_map[pid] for pid in ids_comp if pid in preguntas_map]
                        cantidad = min(conf.num_preguntas, len(preguntas_disponibles))
                        if cantidad > 0:
                            preguntas_seleccionadas.extend(random.sample(preguntas_disponibles, cantidad))

            # Fallback: si no hay configuración o no hay categoría, 10 al azar
            if not preguntas_seleccionadas:
                todas = list(Pregunta.objects.filter(exa_cod=examen))
                cantidad = min(10, len(todas))
                preguntas_seleccionadas = random.sample(todas, cantidad) if cantidad > 0 else []

            random.shuffle(preguntas_seleccionadas)

            # Bulk insert de Respuestas (1 operación en lugar de N)
            Respuesta.objects.bulk_create([
                Respuesta(
                    res_cod=f"{intento.int_cod}_{p.pre_cod}",
                    int_cod=intento,
                    pre_cod=p,
                )
                for p in preguntas_seleccionadas
            ])
        
        # Extraer las preguntas desde las Respuestas atadas al Intento
        respuestas = Respuesta.objects.filter(int_cod=intento).select_related('pre_cod', 'opc_cod')
        preguntas_ids = respuestas.values_list('pre_cod', flat=True)
        preguntas = Pregunta.objects.filter(pre_cod__in=preguntas_ids).prefetch_related('opciones')
        
        serializer = PreguntaSerializer(preguntas, many=True, context={'request': request})

        # Respuestas ya guardadas (para resume)
        respuestas_guardadas = []
        for r in respuestas:
            if r.opc_cod_id or r.res_tex:
                respuestas_guardadas.append({
                    'pre_cod': r.pre_cod_id,
                    'opc_cod': r.opc_cod_id,
                    'res_tex': r.res_tex or ''
                })
        
        return Response({
            'int_cod': intento.int_cod,
            'exa_nom': examen.exa_nom,
            'exa_des': examen.exa_des,
            'questions': serializer.data,
            'respuestas': respuestas_guardadas
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def save_progress(self, request, pk=None):
        """Guarda respuesta en modo intento interrumpible"""
        user = request.user
        int_cod = request.data.get('int_cod')
        
        try:
            intento = Intento.objects.get(int_cod=int_cod, usu_cod=user, exa_fec_fin__isnull=True)
        except Intento.DoesNotExist:
            return Response({'error': 'Intento no encontrado o ya finalizado.'}, status=status.HTTP_404_NOT_FOUND)

        respuestas_data = request.data.get('respuestas', [])
        
        for data in respuestas_data:
            pre_cod = data.get('pre_cod')
            opc_cod = data.get('opc_cod')
            res_tex = data.get('res_tex', '')

            try:
                res = Respuesta.objects.get(int_cod=intento, pre_cod=pre_cod)
                if opc_cod:
                    opcion = Opcion.objects.get(opc_cod=opc_cod)
                    res.opc_cod = opcion
                    res.res_cor = opcion.opc_cor
                    res.res_pun = res.pre_cod.pre_pun if opcion.opc_cor else 0
                else:
                    res.res_tex = res_tex
                    # Open ended logica
                    if res_tex.strip():
                        res.res_cor = True
                        res.res_pun = res.pre_cod.pre_pun
                    else:
                        res.res_cor = False
                        res.res_pun = 0

                res.save()
            except (Respuesta.DoesNotExist, Opcion.DoesNotExist):
                continue
                
        return Response({'status': 'Guardado'})

    @action(detail=True, methods=['post'], permission_classes=[IsStaff])
    def bulk_save(self, request, pk=None):
        """
        Guarda todo el examen de golpe: Metadatos, Preguntas y Opciones.
        """
        examen = self.get_object()
        data = request.data

        try:
            # 1. Actualizar Metadatos
            examen.exa_nom = data.get('exa_nom', examen.exa_nom)
            examen.exa_des = data.get('exa_des', examen.exa_des)
            examen.save()

            questions_data = data.get('questions', [])

            with transaction.atomic():
                for q_data in questions_data:
                    pre_cod = q_data.get('pre_cod')
                    is_deleted = q_data.get('isDeleted', False)
                    is_new = q_data.get('isNew', False)

                    if is_deleted:
                        if not is_new:
                            Pregunta.objects.filter(pre_cod=pre_cod, exa_cod=examen).delete()
                        continue

                    # Crear o Actualizar
                    pre_defaults = {
                        'pre_tex': q_data.get('pre_tex', ''),
                        'pre_pun': q_data.get('pre_pun', 10),
                        'pre_tie': q_data.get('pre_tie', 60),
                        'tip_pre_cod_id': q_data.get('tip_pre_cod', 1),
                    }

                    # Solo actualizar pre_fot si viene explícitamente en el JSON y no es un archivo (bulk_save es JSON)
                    if 'pre_fot' in q_data and isinstance(q_data['pre_fot'], str):
                        val = q_data['pre_fot']
                        if val.startswith('http://') or val.startswith('https://'):
                            from urllib.parse import urlparse
                            val = urlparse(val).path

                        if val.startswith('/media/'):
                            val = val[len('/media/'):]

                        pre_defaults['pre_fot'] = val.strip('/')

                    if is_new:
                        import uuid
                        new_cod = f"PRE_{uuid.uuid4().hex[:12].upper()}"
                        pregunta = Pregunta.objects.create(
                            pre_cod=new_cod,
                            exa_cod=examen,
                            **pre_defaults
                        )
                    else:
                        pregunta, _ = Pregunta.objects.update_or_create(
                            pre_cod=pre_cod,
                            exa_cod=examen,
                            defaults=pre_defaults
                        )

                    # Procesar Competencia de la Pregunta
                    com_cod = q_data.get('com_cod')
                    if com_cod:
                        try:
                            competencia = Competencia.objects.get(com_cod=com_cod)
                            PreguntaCompetencia.objects.update_or_create(
                                pre_cod=pregunta,
                                defaults={'com_cod': competencia}
                            )
                        except Competencia.DoesNotExist:
                            pass
                    else:
                        PreguntaCompetencia.objects.filter(pre_cod=pregunta).delete()

                    # Procesar Opciones
                    options_data = q_data.get('options', [])
                    for o_data in options_data:
                        opc_cod = o_data.get('opc_cod')
                        o_deleted = o_data.get('isDeleted', False)
                        o_new = o_data.get('isNew', False)

                        if o_deleted:
                            if not o_new:
                                Opcion.objects.filter(opc_cod=opc_cod, pre_cod=pregunta).delete()
                            continue

                        o_defaults = {
                            'opc_tex': o_data.get('opc_tex', ''),
                            'opc_cor': o_data.get('opc_cor', False),
                        }

                        if o_new:
                            import uuid
                            Opcion.objects.create(
                                opc_cod=f"OPC_{uuid.uuid4().hex[:15].upper()}",
                                pre_cod=pregunta,
                                **o_defaults
                            )
                        else:
                            Opcion.objects.update_or_create(
                                opc_cod=opc_cod,
                                pre_cod=pregunta,
                                defaults=o_defaults
                            )

                # --- Procesar Configuración de Examen (Categorías y Competencias) ---
                config_data = data.get('config', [])
                if config_data is not None:
                    # 1. Limpiar configuración actual
                    CategoriaExamen.objects.filter(exa_cod=examen).delete()
                    ExamenCategoriaCompetencia.objects.filter(exa_cod=examen).delete()

                    # 2. Recrear según la nueva data
                    for c_item in config_data:
                        cat_cod = c_item.get('cat_cod')
                        if not cat_cod: continue

                        try:
                            categoria = Categoria.objects.get(cat_cod=cat_cod)
                            # Vincular categoría al examen
                            CategoriaExamen.objects.get_or_create(exa_cod=examen, cat_cod=categoria)

                            # Vincular competencias de esta categoría
                            comp_configs = c_item.get('competencies', [])
                            for co_conf in comp_configs:
                                com_cod = co_conf.get('com_cod')
                                num_preg = co_conf.get('num_preguntas', 0)
                                if not com_cod: continue

                                try:
                                    competencia = Competencia.objects.get(com_cod=com_cod)
                                    ExamenCategoriaCompetencia.objects.create(
                                        exa_cod=examen,
                                        cat_cod=categoria,
                                        com_cod=competencia,
                                        num_preguntas=num_preg
                                    )
                                except Competencia.DoesNotExist:
                                    continue
                        except Categoria.DoesNotExist:
                            continue

            # Aseguramos devolver todo el estado fresco para sincronizar el frontend
            updated_questions = Pregunta.objects.filter(exa_cod=examen).prefetch_related('opciones')

            return Response({
                'status': 'ok',
                'exam': ExamenSerializer(examen, context={'request': request}).data,
                'questions': PreguntaSerializer(updated_questions, many=True, context={'request': request}).data
            })

        except Exception as e:
            import traceback
            return Response(
                {'error': str(e), 'detail': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def get_config(self, request, pk=None):
        examen = self.get_object()
        
        # Obtener categorías asociadas
        categorias_ids = CategoriaExamen.objects.filter(exa_cod=examen).values_list('cat_cod', flat=True)
        
        config = []
        for cat_cod in categorias_ids:
            competencias = ExamenCategoriaCompetencia.objects.filter(exa_cod=examen, cat_cod=cat_cod)
            config.append({
                'cat_cod': cat_cod,
                'competencies': [
                    {
                        'com_cod': comp.com_cod_id,
                        'num_preguntas': comp.num_preguntas
                    } for comp in competencias
                ]
            })
            
        return Response(config)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def finish_attempt(self, request, pk=None):
        MAX_SCORED_ATTEMPTS = 3
        user = request.user
        int_cod = request.data.get('int_cod')
        
        try:
            intento = Intento.objects.get(int_cod=int_cod, usu_cod=user, exa_fec_fin__isnull=True)
        except Intento.DoesNotExist:
            return Response({'error': 'Intento no encontrado o ya finalizado.'}, status=status.HTTP_400_BAD_REQUEST)

        examen = intento.exa_cod

        # Calcular puntos totales y correctas desde Respuestas
        respuestas = Respuesta.objects.filter(int_cod=intento).select_related('pre_cod')
        total_puntos = respuestas.aggregate(models.Sum('res_pun'))['res_pun__sum'] or 0
        correct_count = respuestas.filter(res_cor=True).count()
        total_questions = respuestas.count()

        intento.exa_pun_tot = total_puntos
        intento.exa_fec_fin = timezone.now()
        intento.save()

        # Contar intentos COMPLETADOS (incluyendo el actual)
        intentos_completados = Intento.objects.filter(
            usu_cod=user, exa_cod=examen, exa_fec_fin__isnull=False
        ).count()
        attempts_left = max(0, MAX_SCORED_ATTEMPTS - intentos_completados)
        counts_for_score = intentos_completados <= MAX_SCORED_ATTEMPTS

        # Actualizar usu_pun_tot solo si este intento cuenta y mejoró el mejor puntaje previo
        if counts_for_score:
            mejor_puntaje_previo = Intento.objects.filter(
                usu_cod=user, exa_cod=examen, exa_fec_fin__isnull=False
            ).exclude(int_cod=intento.int_cod).aggregate(models.Max('exa_pun_tot'))['exa_pun_tot__max'] or 0

            mejora = max(0, total_puntos - mejor_puntaje_previo)
            if mejora > 0:
                user.usu_pun_tot = (user.usu_pun_tot or 0) + mejora
                user.save()

        return Response({
            'status': 'finalizado',
            'score': total_puntos,
            'correct_count': correct_count,
            'total_questions': total_questions,
            'counts_for_score': counts_for_score,
            'attempts_left': attempts_left,
            'total_user_score': user.usu_pun_tot,
        })


class IntentoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Intento.objects.all()
    serializer_class = IntentoSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return Intento.objects.all()
        return Intento.objects.filter(usu_cod=self.request.user)

class PreguntaViewSet(viewsets.ModelViewSet):
    queryset = Pregunta.objects.all()
    serializer_class = PreguntaSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    permission_classes = [IsStaff]

    def get_queryset(self):
        exa_cod = self.request.query_params.get('exa_cod')
        if exa_cod:
            return Pregunta.objects.filter(exa_cod=exa_cod)
        return super().get_queryset()

    def perform_create(self, serializer):
        # Inyectar exa_cod desde query params si no viene en el body
        exa_cod = self.request.data.get('exa_cod')
        if not exa_cod:
            exa_cod = self.request.query_params.get('exa_cod')
        
        if exa_cod:
            try:
                examen = Examen.objects.get(exa_cod=exa_cod)
                serializer.save(exa_cod=examen)
            except Examen.DoesNotExist:
                serializer.save() # Fallback, let serializer handle error if exa_cod is missing and required
        else:
            serializer.save()

class OpcionViewSet(viewsets.ModelViewSet):
    queryset = Opcion.objects.all()
    serializer_class = OpcionSerializer
    permission_classes = [IsStaff]

    def perform_create(self, serializer):
        # Inyectar pre_cod desde data o query params
        pre_cod = self.request.data.get('pre_cod')
        if not pre_cod:
            pre_cod = self.request.query_params.get('pre_cod')
        
        if pre_cod:
            try:
                pregunta = Pregunta.objects.get(pre_cod=pre_cod)
                serializer.save(pre_cod=pregunta)
            except Pregunta.DoesNotExist:
                serializer.save()
        else:
            serializer.save()
class TipoPreguntaViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = TipoPregunta.objects.all().order_by('tip_pre_cod')
    serializer_class = TipoPreguntaSerializer
    permission_classes = [IsAuthenticated]
